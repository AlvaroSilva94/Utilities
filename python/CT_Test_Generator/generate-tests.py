#-------------------------------------------------------------------------------------------
# Purpose: Generation of test case scripts with the complete structure, ready
# to implement the test logic, from the Doors export. This 
# script generates .c files for the specified test.
#
# Usage example:
#   python generate-tests.py
#   - the script will then prompt the user to give as input:
#       - The number of tests to be created;
#       - The requirement(s) to be tested;
#       - The path to the DOORs export file with .xlsx extension;
# Notes:
#   The script expects the following files and directories to be present:
#       \output\                             -> folder to save the generated scripts
#       \template\cmocka-template.c          -> template for the .c files      
#
# - Tested in Windows environment with Python 3.9.10 
#
# Modification History:
# CR        Date           Name           Comment
# --------  ------------   ------------   ------------------------------------------------------
# 0000      02/12/2022     Alvaro Silva   Initial version
# 0001      03/12/2022     Alvaro Silva   Added possibility to select multiple requirements
#                                         and create multiple test cases at once
# 0002      13/12/2022     Alvaro Silva   Added automatic date retrival and name input
# 0003      14/12/2022     Alvaro Silva   Added retrieval of data from plan.md
# 0004      14/12/2022     Alvaro Silva   Added config file to store name, export and plan paths
# 0005      22/12/2022     Alvaro Silva   Added option to never use plan if not needed
# 0006      22/12/2022     Alvaro Silva   Fixed dependance of fixed column structure
#-----------------------------------------------------------------------------------------------

from msilib.schema import ComboBox
import openpyxl
import os
import glob
import itertools 
from collections import namedtuple
import re
import csv
from datetime import datetime

# Template file paths
CMOCKA_TEMPLATE_NAME = "template\cmocka-template.c"
OUTPUT_DIR = "output\\"
PLAN_MD = ""
NAME_ME = ""
DOORS_EXPORT_PATH = ""
INIT = "init\config.csv"

# Getting today's date
date_string = datetime.today().strftime('%d/%m/%Y')

# Getting the req as user input 
TC_NUM_STR = input("Please input the number of tests to be created: ") 
TC_NUM = int(TC_NUM_STR)

# Creating a list to get all user input values
FIND_VALUE_NUM = []

for x in range(TC_NUM):
    TEMP_NUM = input("Please input the desired requirement to be tested: ")
    FIND_VALUE_NUM.append(TEMP_NUM)

#Allowing to never use plan for TC creation
while True:
    TC_OPT_PLAN = input("Do you want to use info from the plan?(Y/N) ")
    if TC_OPT_PLAN not in ('Y', 'N'):
        print("Please select \"Y\" or \"N\".")
    else:
        break

# Creating lists to store RequirementID, function to be tested and file name
FIND_VALUE = []
C_FUNCT_NAME = []
TEST_FILE_NAME = []

for y in FIND_VALUE_NUM:
    BASE = "E600SWRS-"
    FIND_VALUE_TEMP = "".join((BASE, y))
    FIND_VALUE.append(FIND_VALUE_TEMP)  

    C_BASE_FUNCT_NAME = "CT_GSTMCTRL_etcs_swrs_"
    C_FUNCT_NAME.append("".join((C_BASE_FUNCT_NAME,y)))

    TEST_FILE_PREFIX = "ct_gstmctrl_etcs_swrs_"
    TEST_FILE_POSTFIX = ".c"

    TEST_FILE_PREFIX = "".join((TEST_FILE_PREFIX,y))
    TEST_FILE_NAME.append("".join((TEST_FILE_PREFIX,TEST_FILE_POSTFIX)))

def InitalConfig(csv_file):

    line_number = 1 
    global NAME_ME, DOORS_EXPORT_PATH, PLAN_MD     
    #Read from csv, [line][column]
    with open(csv_file, 'r') as config:
        init_csv = csv.reader(config)
        init_csv = list(init_csv)
        NAME_ME = init_csv[line_number][0]
        DOORS_EXPORT_PATH = init_csv[line_number][1]
        PLAN_MD = init_csv[line_number][2]

#Function to get values from excel and iterate through lists 
def getValuesFromExcelFile(file_path):
    file_path = os.path.abspath(file_path)
    tg_xlsx = openpyxl.load_workbook(file_path)

    #Dictionary to store test cases
    test_cases_dict={}

    for (id, file, reqID) in itertools.zip_longest(C_FUNCT_NAME, TEST_FILE_NAME, FIND_VALUE):

        #Values to be inserted in the tuple
        test_case_id = id
        fileNameC = file
        requirementID = reqID
        requirementTxt = ""
        coverage = ""
        componentName = ""
        functionFromReq = ""

        #creating a tuple to store values
        test_case = namedtuple('test_case', ['test_case_id', 'componentName', 'fileNameC','coverage','requirementID', 'requirementTxt', 'functionFromReq'])

        #get data from test cases
        sheet_data = tg_xlsx['RichText']

        #Creating a dictionary to store all column names
        Export_col_names = {}
        curr_col = 0
        #Iterate through all columns and save the column value
        for column in sheet_data.iter_cols(1, sheet_data.max_column):
            Export_col_names[column[0].value] = curr_col
            curr_col +=1
        
        #Intermediate variables to store column number
        RinSW = Export_col_names.get('Realization_in_SW')
        Cov = Export_col_names.get('CoverageStatus')
        GermanCrap = Export_col_names.get('SW-Anforderungs-Spezifikation')
        SubSyst = Export_col_names.get('Subsystem')

        for row in sheet_data.iter_rows():
            for cell in row:
                if cell.value == requirementID:
                    #Get requirement and function to test
                    functionFromReq = sheet_data.cell(row=cell.row, column=RinSW+1).value #Realization_in_SW
                    requirementTxt = sheet_data.cell(row=cell.row, column=GermanCrap+1).value #SW-Anforderungs-Spezifikation 
                    coverage = sheet_data.cell(row=cell.row, column=Cov+1).value #CoverageStatus
                    componentName = sheet_data.cell(row=cell.row, column=SubSyst+1).value #Subsystem

        test_cases_dict[test_case] = test_case(test_case_id, componentName, fileNameC, coverage, requirementID, requirementTxt, functionFromReq)

    return test_cases_dict

#create the .c file for each test case
def create_test_cmocka_template_c(plan, template_path, test_cases_dict):
        
    for test_case in test_cases_dict.values():

        #open template in read only mode, preserving LF
        with open(template_path, "r", newline='') as c_template_file:
            c_template_data = c_template_file.read()

            #replacing in the template file to create the new file
            c_new_file = re.sub('@fileNameC@', test_case.fileNameC, c_template_data)
            c_new_file = re.sub('@test_case_id@', test_case.test_case_id, c_new_file)
            c_new_file = re.sub('@componentName@',test_case.componentName, c_new_file)
            c_new_file = re.sub('@coverage@', test_case.coverage, c_new_file)
            c_new_file = re.sub('@functionFromReq@', test_case.functionFromReq, c_new_file)
            c_new_file = re.sub('@requirement@', test_case.requirementID, c_new_file)
            c_new_file = re.sub('@authorMe@', NAME_ME, c_new_file)
            c_new_file = re.sub('@dateNow@', date_string, c_new_file)

            if (TC_OPT_PLAN == 'Y'):
                with open(plan) as myplan:
                    content = myplan.read()
                    search_string = r"(?s)(?<=" + re.escape(test_case.requirementID) + r").*?(?=# E)"

                    if re.search(search_string, content, re.DOTALL).group():
                        text = re.search(search_string, content, re.DOTALL).group()
                        c_new_file = re.sub('@plan@', text, c_new_file)

            if (TC_OPT_PLAN == 'N'):
                    c_new_file = re.sub('@requirementTxt@', test_case.requirementTxt, c_new_file)

            #generate .c file
            with open(OUTPUT_DIR + test_case.fileNameC, "w", newline='') as c_file:
                c_file.write(c_new_file)

            print("File created: " + test_case.fileNameC + "!")

def RemoveOldFiles():

    # Search files with .txt extension in current directory
    pattern = "*.c"
    OldTests = glob.glob(pattern)

    # deleting the files with txt extension
    for file in OldTests:
        os.remove(file)


print("Loading initial configuration!\n")
InitalConfig(INIT)

print("Getting data from Doors export!\n")
ct_dict = getValuesFromExcelFile(DOORS_EXPORT_PATH)

print("Removing all files from output folder!\n")
RemoveOldFiles()

print("Generating .c files\n")
create_test_cmocka_template_c(PLAN_MD, CMOCKA_TEMPLATE_NAME, ct_dict)

print("\nAll done!\n")